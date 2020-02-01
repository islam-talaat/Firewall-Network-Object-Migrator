from openpyxl import Workbook, load_workbook
from termcolor import colored
import json
def read_xlsx_file(file_Dir):
    wb = load_workbook(file_Dir)
    print(colored('this file contain the following sheets : ', 'red'))
    for sheet in wb:
        print(colored(sheet.title,'red'))
    sheet_name =input(colored('Enter Your sheet name : ', 'blue'))
    ws = wb[sheet_name]
    n_rules = ws.max_row
    OriginalFormList = []
    for i in range (2, n_rules+1):
        Obj_Name = ws['A' + str(i)].value
        Obj_Value = ws['B' + str(i)].value
        print(Obj_Value)
        Obj_Description = ''
        if "/" in Obj_Value:
            Obj_Type = "Network"
        elif "-" in Obj_Value:
            Obj_Type = "Range"
        else:
            Obj_Type = "Host"
        element = [{"name": Obj_Name, "value": Obj_Value, "type": Obj_Type, "description": Obj_Description}]
        OriginalFormList = OriginalFormList + element
    with open('OriginalFormList.json', 'w') as f:
        json.dump(OriginalFormList, f, indent=4)
        f.write("\n")
    return (OriginalFormList)